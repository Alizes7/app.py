# NF-PDF Excel Extractor v2.3 - Production Ready
# -*- coding: utf-8 -*-
import io, re, logging, traceback, unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, List
import fitz, pdfplumber, pandas as pd, streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("NF-App")

# Regex Patterns
RE_CNPJ = re.compile(r'\b(\d{2}[.\s]?\d{3}[.\s]?\d{3}[\/\s]?\d{4}[-\s]?\d{2})\b')
RE_CPF = re.compile(r'\b(\d{3}[.\s]?\d{3}[.\s]?\d{3}[-\s]?\d{2})\b')
RE_DATA = re.compile(r'\b(\d{2}[\/\-\.]\d{2}[\/\-\.]\d{4})\b')
RE_VALOR = re.compile(r'(?:R?\$[\s]*)?(\d{1,3}(?:\.\d{3})+,\d{2}|\d+,\d{2})(?!\d)')
RE_NUMERO_NOTA = [
    re.compile(r'N[°º#]\s*(\d{4,})', re.IGNORECASE),
    re.compile(r'N[UÚ]MERO\s*:?\s*(\d{4,})', re.IGNORECASE),
    re.compile(r'NF[SE]?[ -]*(\d{4,})', re.IGNORECASE),
    re.compile(r'NOTA\s*FISCAL[:\s]*(\d{4,})', re.IGNORECASE),
    re.compile(r'RPS[:\s]*(\d{4,})', re.IGNORECASE),
]

def normalizar(texto: str) -> str:
    if not texto: return ""
    return "".join(c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c)).lower().strip()

def limpar_valor(valor_str: str) -> Optional[float]:
    try:
        limpo = re.sub(r"[R$\s]", "", str(valor_str)).replace(".", "").replace(",", ".")
        return float(limpo)
    except:
        return None

def formatar_cnpj(cnpj_str: str) -> str:
    digitos = re.sub(r"\D", "", cnpj_str)
    if len(digitos) == 14:
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    return cnpj_str

def extrair_cnpjs(texto: str) -> List[str]:
    return list(dict.fromkeys([formatar_cnpj(c) for c in RE_CNPJ.findall(texto) + RE_CPF.findall(texto)]))

def extrair_numero_nota(texto: str) -> Optional[str]:
    for padrao in RE_NUMERO_NOTA:
        m = padrao.search(texto)
        if m:
            numero = re.sub(r"\D", "", m.group(1))
            if 4 <= len(numero) <= 15 and len(numero) not in [11, 14]:
                return numero
    return None

def extrair_data(texto: str) -> Optional[str]:
    m = RE_DATA.search(texto)
    return m.group(1).replace("-", "/").replace(".", "/") if m else None

def extrair_valor(texto: str, keywords: List[str]) -> Optional[str]:
    for kw in keywords:
        pos = normalizar(texto).find(normalizar(kw))
        if pos != -1:
            trecho = texto[pos + len(kw):pos + len(kw) + 80]
            matches = RE_VALOR.findall(trecho)
            if matches:
                return max(matches, key=lambda x: limpar_valor(x) or 0)
    return None

def processar_pdf(pdf_bytes: bytes, nome: str) -> Dict[str, Any]:
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    try:
        # Extracao com PyMuPDF
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        texto_fitz = "\n".join([page.get_text() for page in doc])
        doc.close()
        
        # Extracao com pdfplumber (melhor para tabelas)
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto_plumber = "\n".join([p.extract_text() or "" for p in pdf.pages])
        
        texto = texto_plumber if len(texto_plumber) > len(texto_fitz) else texto_fitz
        if not texto.strip():
            return {"Arquivo": nome, "Processado Em": timestamp, "Erro": "PDF sem texto - pode ser imagem"}
        
        cnpjs = extrair_cnpjs(texto)
        dados = {
            "Arquivo": nome,
            "Processado Em": timestamp,
            "Número da Nota": extrair_numero_nota(texto) or "",
            "Data de Emissão": extrair_data(texto) or "",
            "CNPJ Prestador": cnpjs[0] if len(cnpjs) > 0 else "",
            "CNPJ Tomador": cnpjs[1] if len(cnpjs) > 1 else "",
            "Valor Bruto (R$)": extrair_valor(texto, ["valor total", "valor dos servicos", "valor bruto", "total"]) or "",
            "Valor Líquido (R$)": extrair_valor(texto, ["valor liquido", "valor a receber", "liquido"]) or "",
            "Base de Cálculo (R$)": extrair_valor(texto, ["base de calculo", "base calc"]) or "",
            "Valor ISS (R$)": extrair_valor(texto, ["valor do iss", "iss calculado", "iss:"]) or "",
            "Erro": ""
        }
        
        if not dados["Número da Nota"] or not dados["CNPJ Prestador"]:
            dados["Erro"] = "Campos criticos nao encontrados"
        
        return dados
        
    except Exception as e:
        return {"Arquivo": nome, "Processado Em": timestamp, "Erro": str(e)}

def gerar_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Notas Fiscais"
    ws.freeze_panes = "A2"
    
    headers = list(df.columns)
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")
    
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        ws.append(list(row))
        for c_idx, cell in enumerate(ws[r_idx], 1):
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            )
            if c_idx == len(headers) and cell.value:
                cell.fill = PatternFill("solid", fgColor="FDECEA")
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
    
    wb.save(output)
    return output.getvalue()

# Streamlit UI
st.set_page_config(page_title="NF-PDF para Excel", page_icon="📄", layout="wide")

st.markdown("""
<style>
    .main {background-color: #f5f7fa;}
    .stButton>button {background-color: #2E86AB; color: white; font-weight: bold; border-radius: 8px; padding: 0.5rem 2rem;}
    .stButton>button:hover {background-color: #1E3A5F;}
    h1 {color: #1E3A5F; font-family: 'Segoe UI', sans-serif;}
    .info-box {background: #e8f4f8; padding: 1rem; border-radius: 8px; border-left: 4px solid #2E86AB;}
</style>
""", unsafe_allow_html=True)

st.title("📄 Extrator de Notas Fiscais")
st.markdown("<div class='info-box'><b>Como usar:</b> Envie os PDFs das notas fiscais abaixo e clique em Processar. O sistema extrai numero, data, CNPJs e valores automaticamente.</div>", unsafe_allow_html=True)

uploaded = st.file_uploader("Selecione os PDFs", type=["pdf"], accept_multiple_files=True)

if uploaded:
    if st.button("🚀 Processar Notas", use_container_width=True):
        progress = st.progress(0)
        status = st.empty()
        resultados = []
        
        for i, arquivo in enumerate(uploaded):
            status.info(f"Processando {arquivo.name}... ({i+1}/{len(uploaded)})")
            dados = processar_pdf(arquivo.read(), arquivo.name)
            resultados.append(dados)
            progress.progress((i + 1) / len(uploaded))
        
        progress.empty()
        status.empty()
        
        df = pd.DataFrame(resultados)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Total", len(df))
        col2.metric("Sucesso", len(df[df["Erro"] == ""]))
        col3.metric("Com Erro", len(df[df["Erro"] != ""]))
        
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        excel = gerar_excel(df)
        st.download_button(
            label="📥 Baixar Excel",
            data=excel,
            file_name=f"Notas_Fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        if any(df["Erro"] != ""):
            st.warning("⚠️ Alguns arquivos tiveram problemas. Verifique a coluna 'Erro' na tabela acima.")
